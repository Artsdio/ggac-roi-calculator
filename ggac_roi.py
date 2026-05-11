import streamlit as st
import plotly.graph_objects as go
import pandas as pd
import openpyxl
import io
from pdf_report import generate_pdf

st.set_page_config(
    page_title="GGAC ROI Calculator",
    page_icon="💪",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Plus+Jakarta+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'Plus Jakarta Sans', sans-serif; }
.main { background-color: #F7F6F2; }
.block-container { padding-top: 2rem; padding-bottom: 2rem; }
.metric-card {
    background: #FFFFFF; border-radius: 16px;
    padding: 1.25rem 1.5rem; border: 1px solid #E8E6DE; margin-bottom: 12px;
}
.metric-label { font-size: 12px; color: #888780; font-weight: 500; text-transform: uppercase; letter-spacing: .06em; margin-bottom: 4px; }
.metric-value { font-size: 22px; font-weight: 700; font-family: 'DM Mono', monospace; margin-bottom: 2px; }
.metric-sub { font-size: 12px; color: #888780; }
.green { color: #0F6E56; } .blue { color: #185FA5; } .amber { color: #854F0B; } .red { color: #A32D2D; }
.section-header {
    font-size: 11px; font-weight: 600; color: #888780; text-transform: uppercase; letter-spacing: .1em;
    border-bottom: 1px solid #E8E6DE; padding-bottom: 8px; margin-bottom: 16px; margin-top: 24px;
}
.breakdown-table { width: 100%; border-collapse: collapse; }
.breakdown-table td { padding: 8px 0; border-bottom: 1px solid #F0EEE8; font-size: 14px; }
.breakdown-table td:last-child { text-align: right; font-family: 'DM Mono', monospace; font-weight: 500; }
.breakdown-table .total-row td { font-weight: 700; font-size: 15px; border-top: 2px solid #D3D1C7; border-bottom: none; }
.breakdown-table .sub-row td { font-size: 12px; color: #888780; padding: 3px 0 3px 16px; border-bottom: none; font-style: italic; }
.badge { display: inline-block; padding: 3px 10px; border-radius: 20px; font-size: 11px; font-weight: 600; }
.badge-green { background: #E1F5EE; color: #0F6E56; }
.badge-red   { background: #FCEBEB; color: #A32D2D; }
.badge-amber { background: #FAEEDA; color: #854F0B; }
.written-card {
    background: #FFFFFF; border-radius: 16px; padding: 1.5rem 2rem;
    border: 1px solid #E8E6DE; font-size: 14px; line-height: 1.8; color: #444441;
}
.written-card h4 { color: #2C2C2A; font-size: 15px; margin-bottom: 8px; margin-top: 16px; }
.written-card h4:first-child { margin-top: 0; }
.hero-banner {
    background: linear-gradient(135deg, #0F6E56 0%, #185FA5 100%);
    border-radius: 20px; padding: 2rem 2.5rem; margin-bottom: 2rem; color: white;
}
.hero-banner h1 { font-size: 28px; font-weight: 700; margin-bottom: 4px; }
.hero-banner p  { font-size: 14px; opacity: .8; margin: 0; }
.skenario-card {
    border-radius: 16px; padding: 1.25rem 1.5rem; margin-bottom: 8px; border: 2px solid transparent;
}
.skenario-pesimis  { background: #FFF5F5; border-color: #FACACA; }
.skenario-moderat  { background: #FFFBF0; border-color: #F5E3A0; }
.skenario-optimis  { background: #F0FBF6; border-color: #A8E6CC; }
.skenario-title    { font-size: 15px; font-weight: 700; margin-bottom: 4px; }
.tab-desc          { font-size: 12px; color: #888780; margin-bottom: 12px; line-height: 1.5; }
</style>
""", unsafe_allow_html=True)

def fmt_rp(n):
    if abs(n) >= 1_000_000:
        return f"Rp {n/1_000_000:.1f}jt"
    return f"Rp {int(round(n)):,}".replace(",", ".")

def fmt_full(n):
    return f"Rp {int(round(abs(n))):,}".replace(",", ".")

CHART_FONT = dict(family="Plus Jakarta Sans", size=12, color="#1a1a1a")
TICK_FONT  = dict(color="#1a1a1a", size=11)
TITLE_FONT = dict(color="#1a1a1a", size=12)


def hitung(u_pct, growth_pct, cost_growth_pct,
           q_private, q_couple, q_group, q_chair, q_tambahan, q_mandiri,
           harga_private, harga_couple, harga_group, harga_chair, harga_tambahan, harga_mandiri,
           peserta_couple, peserta_group, peserta_chair, peserta_tambahan,
           ins_private, ins_couple, ins_group, ins_chair, ins_tambahan, ins_mandiri,
           b_trainer, b_admin, b_mkt, b_maint, b_sdm, b_supply, b_mgmt,
           investasi, modal_pengelola, omzet_min, share_i, share_ii, proj_years):

    u = u_pct / 100
    s_private  = q_private  * u
    s_couple   = q_couple   * u
    s_group    = q_group    * u
    s_chair    = q_chair    * u
    s_tambahan = q_tambahan * u
    s_mandiri  = q_mandiri  * u

    rev_private  = s_private  * harga_private
    rev_couple   = s_couple   * peserta_couple   * harga_couple
    rev_group    = s_group    * peserta_group    * harga_group
    rev_chair    = s_chair    * peserta_chair    * harga_chair
    rev_tambahan = s_tambahan * peserta_tambahan * harga_tambahan
    rev_mandiri  = s_mandiri  * harga_mandiri
    total_rev    = rev_private + rev_couple + rev_group + rev_chair + rev_tambahan + rev_mandiri

    ins_tot_private  = s_private  * ins_private
    ins_tot_couple   = s_couple   * ins_couple
    ins_tot_group    = s_group    * ins_group
    ins_tot_chair    = s_chair    * ins_chair
    ins_tot_tambahan = s_tambahan * ins_tambahan
    ins_tot_mandiri  = s_mandiri  * ins_mandiri
    total_insentif   = (ins_tot_private + ins_tot_couple + ins_tot_group
                        + ins_tot_chair + ins_tot_tambahan + ins_tot_mandiri)

    b_fixed    = b_trainer + b_admin + b_mkt + b_maint + b_sdm + b_supply + b_mgmt
    total_opex = b_fixed + total_insentif

    laba_bersih = total_rev - total_opex
    hasil_ii    = laba_bersih * (share_ii / 100) if laba_bersih > 0 else 0
    hasil_i     = laba_bersih * (share_i  / 100) if laba_bersih > 0 else 0
    bep_met     = total_rev >= omzet_min
    roi_pct     = (hasil_i  * 12 / investasi       * 100) if investasi       > 0 else 0
    roi_ii_pct  = (hasil_ii * 12 / modal_pengelola * 100) if modal_pengelola > 0 else 0
    margin      = (laba_bersih / total_rev * 100) if total_rev > 0 else 0
    total_modal = investasi + modal_pengelola
    pct_modal_i  = investasi       / total_modal * 100 if total_modal > 0 else 0
    pct_modal_ii = modal_pengelola / total_modal * 100 if total_modal > 0 else 0

    months        = proj_years * 12
    cum_owner     = [-investasi]
    cum_pengelola = [-modal_pengelola]
    bep_month     = None
    bep_month_ii  = None

    for m in range(1, months + 1):
        yr      = (m - 1) // 12
        rev_m   = total_rev  * ((1 + growth_pct      / 100) ** yr)
        opex_m  = total_opex * ((1 + cost_growth_pct / 100) ** yr)
        laba_m  = rev_m - opex_m
        earn_i  = laba_m * (share_i  / 100) if laba_m > 0 and rev_m >= omzet_min else 0
        earn_ii = laba_m * (share_ii / 100) if laba_m > 0 and rev_m >= omzet_min else 0
        cum_owner.append(cum_owner[-1] + earn_i)
        cum_pengelola.append(cum_pengelola[-1] + earn_ii)
        if bep_month    is None and cum_owner[-1]     >= 0: bep_month    = m
        if bep_month_ii is None and cum_pengelola[-1] >= 0: bep_month_ii = m

    return dict(
        total_rev=total_rev, total_opex=total_opex, laba_bersih=laba_bersih,
        hasil_ii=hasil_ii, hasil_i=hasil_i, bep_met=bep_met,
        roi_pct=roi_pct, roi_ii_pct=roi_ii_pct,
        margin=margin, cum_owner=cum_owner, cum_pengelola=cum_pengelola,
        bep_month=bep_month, bep_month_ii=bep_month_ii, months=months,
        modal_pengelola=modal_pengelola, investasi=investasi,
        total_modal=total_modal, pct_modal_i=pct_modal_i, pct_modal_ii=pct_modal_ii,
        rev_private=rev_private, rev_couple=rev_couple, rev_group=rev_group,
        rev_chair=rev_chair, rev_tambahan=rev_tambahan, rev_mandiri=rev_mandiri,
        total_insentif=total_insentif,
        ins_tot_private=ins_tot_private, ins_tot_couple=ins_tot_couple,
        ins_tot_group=ins_tot_group, ins_tot_chair=ins_tot_chair,
        ins_tot_tambahan=ins_tot_tambahan, ins_tot_mandiri=ins_tot_mandiri,
        b_trainer=b_trainer, b_admin=b_admin, b_mkt=b_mkt, b_maint=b_maint,
        b_sdm=b_sdm, b_supply=b_supply, b_mgmt=b_mgmt,
        s_private=s_private, s_couple=s_couple, s_group=s_group,
        s_chair=s_chair, s_tambahan=s_tambahan, s_mandiri=s_mandiri,
    )


# ─── HERO ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="hero-banner">
  <h1>💪 GGAC ROI Calculator</h1>
  <p>GREKO Golden Age Center — Simulasi Investasi &amp; Proyeksi Keuangan</p>
</div>
""", unsafe_allow_html=True)


# ─── HELPER FORMAT RIBUAN (titik, gaya Indonesia) ────────────────────────────
def rp(n):
    """Format angka dengan pemisah ribuan titik: 1.500.000"""
    return f"{int(round(n)):,}".replace(",", ".")

def rp_cap(n):
    """Caption berwarna hijau tua dengan format Rp x.xxx.xxx"""
    return f'<span style="font-size:12px;color:#0F6E56;font-weight:600">→ Rp {rp(n)}</span>'


# ─── PARSER EXCEL MODAL_GGAC ────────────────────────────────────────────────
def parse_modal_excel(file_bytes):
    """
    Baca Modal_GGAC_v2 atau v3 — scan dinamis, tidak bergantung posisi baris.
    Deteksi section dari header teks, hitung total dari item langsung.
    """
    wb   = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    hasil = {}

    # ── Sheet 1: Modal Pengelola ──────────────────────────────────────────────
    if "Modal Pengelola" in wb.sheetnames:
        ws = wb["Modal Pengelola"]
        for row in ws.iter_rows():
            v0 = str(row[0].value or "").strip().upper()
            if "TOTAL" in v0:
                v = row[6].value
                if isinstance(v, (int, float)):
                    hasil["modal_pengelola"] = int(round(v))
                break
        items = []
        for row in ws.iter_rows(min_row=14):
            if not isinstance(row[0].value, int):
                continue
            nb = row[6].value
            items.append({
                "no":    row[0].value,
                "nama":  str(row[1].value or ""),
                "harga": int(row[2].value) if isinstance(row[2].value, (int,float)) else 0,
                "thn":   int(row[3].value) if isinstance(row[3].value, (int,float)) else 2020,
                "nb":    int(round(nb))    if isinstance(nb, (int,float)) else 0,
            })
        hasil["peralatan"] = items

    # ── Sheet 2: Investasi Investor — scan dinamis ────────────────────────────
    if "Investasi Investor" in wb.sheetnames:
        ws2 = wb["Investasi Investor"]

        fase1_gym, fase2_gym, pend_items = [], [], []
        section   = None
        buf_pct   = 0.10

        for row in ws2.iter_rows():
            a = str(row[0].value or "").strip().upper()
            b = str(row[1].value or "").strip().upper()

            # ── Deteksi section header ────────────────────────────────────────
            # Cek FASE 2 dulu (sebelum FASE 1) karena header Fase 2 bisa
            # mengandung teks "ROI Fase 1" di keterangannya
            if ("FASE 2" in a or "FASE 2" in b) and "SUBTOTAL" not in a:
                section = "fase2"; continue
            if ("FASE 1" in a or "FASE 1" in b) and "SUBTOTAL" not in a \
                    and "FASE 2" not in a:
                section = "fase1"; continue
            if ("PENUNJANG" in a or "PENUNJANG" in b) and "SUBTOTAL" not in a \
                    and "FASE" not in a:
                section = "pend"; continue
            # Stop saat masuk area Ringkasan — bukan lagi area item
            if any(kw in a for kw in ["RINGKASAN","PERHITUNGAN","KOMPONEN"]):
                break

            # ── Buffer % — ambil dari mana saja yang masuk akal ──────────────
            if "BUFFER" in a:
                for ci in [2, 3, 4]:          # cek col C, D, E
                    v = row[ci].value if ci < len(row) else None
                    if isinstance(v, float) and 0 < v <= 1:
                        buf_pct = float(v)
                        break
                continue

            # ── Lewati baris subtotal / total / header ────────────────────────
            if any(kw in a for kw in ["SUBTOTAL","TOTAL","NO.","NAMA","ITEM","⚡","HALAMAN"]):
                continue

            # ── Baris item: col A integer DAN col B harus ada nama ───────────
            if not isinstance(row[0].value, int):
                continue
            if section is None:
                continue

            nama = row[1].value
            qty  = row[2].value
            hrg  = row[3].value

            # Skip baris yang tidak punya nama (misal baris kosong bernomor)
            if not nama or str(nama).strip() == "":
                continue
            # Skip baris header yang kadang bernomor
            if str(nama).strip().upper() in ["NAMA PERALATAN","ITEM FASILITAS","NAMA","KETERANGAN"]:
                continue

            q = int(qty) if isinstance(qty, (int,float)) else 0
            h = int(hrg) if isinstance(hrg, (int,float)) else 0
            entry = {
                "nama":  str(nama).strip(),
                "qty":   q,
                "harga": h,
                "total": q * h,
            }
            if   section == "fase1": fase1_gym.append(entry)
            elif section == "fase2": fase2_gym.append(entry)
            elif section == "pend":  pend_items.append(entry)

        # ── Hitung subtotal dari item (bukan dari formula Excel) ──────────────
        sub_f1   = sum(i["total"] for i in fase1_gym  if i["qty"] > 0)
        sub_f2   = sum(i["total"] for i in fase2_gym  if i["qty"] > 0)
        sub_pend = sum(i["total"] for i in pend_items if i["qty"] > 0)

        buf_f1   = int(round((sub_f1 + sub_pend)            * buf_pct))
        buf_f12  = int(round((sub_f1 + sub_f2 + sub_pend)   * buf_pct))
        total_f1  = sub_f1 + sub_pend + buf_f1
        total_f12 = sub_f1 + sub_f2 + sub_pend + buf_f12

        hasil.update({
            "sub_fase1":     sub_f1,
            "sub_fase2":     sub_f2,
            "sub_pend":      sub_pend,
            "buffer_pct":    buf_pct,
            "investasi_f1":  total_f1,
            "investasi_f12": total_f12,
            "investasi":     total_f1,
        })

        hasil["fase1_gym"]   = fase1_gym
        hasil["fase2_gym"]   = fase2_gym
        hasil["pend_items"]  = pend_items
        hasil["fase1_aktif"] = [i for i in fase1_gym  if i["qty"] > 0]
        hasil["fase2_aktif"] = [i for i in fase2_gym  if i["qty"] > 0]
        hasil["pend_aktif"]  = [i for i in pend_items if i["qty"] > 0]

    return hasil


# ─── FILE UPLOADER ───────────────────────────────────────────────────────────
with st.expander("📂 Import dari Excel — Modal_GGAC_v3.xlsx", expanded=False):
    st.markdown(
        "Upload file **Modal_GGAC_v3.xlsx** untuk mengisi otomatis nilai investasi & "
        "modal pengelola di sidebar. Ubah angka di Excel → simpan → upload ulang → "
        "semua input terupdate."
    )
    uploaded = st.file_uploader(
        "Pilih file Modal_GGAC_v3.xlsx", type=["xlsx"],
        key="modal_upload", label_visibility="collapsed"
    )

    xl_data = {}
    if uploaded is not None:
        try:
            xl_data = parse_modal_excel(uploaded.read())

            # ── Pilih fase ───────────────────────────────────────────────────
            fase_pilihan = st.radio(
                "Hitung ROI untuk fase mana?",
                ["Fase 1 saja (operasional sekarang)",
                 "Fase 1 + Fase 2 (setelah scale-up)"],
                horizontal=True, key="fase_radio"
            )
            inv_aktif = (xl_data.get("investasi_f1", 0)
                         if "Fase 1 saja" in fase_pilihan
                         else xl_data.get("investasi_f12", 0))
            xl_data["investasi"]  = inv_aktif
            xl_data["fase_label"] = ("Fase 1 — Prioritas/Urgent"
                                     if "Fase 1 saja" in fase_pilihan
                                     else "Fase 1 + Fase 2 — Termasuk Pengembangan")

            # ── Ringkasan modal ───────────────────────────────────────────────
            f1  = xl_data.get("sub_fase1", 0)
            f2  = xl_data.get("sub_fase2", 0)
            pen = xl_data.get("sub_pend", 0)
            buf_pct = xl_data.get("buffer_pct", 0.1)
            mp  = xl_data.get("modal_pengelola", 0)

            col_s1, col_s2, col_s3 = st.columns(3)
            with col_s1:
                st.markdown(f"""
                <div style="background:#E8F5EE;border-radius:10px;padding:.75rem 1rem;
                            border-left:4px solid #0F6E56">
                  <div style="font-size:11px;color:#888780;font-weight:600;
                              text-transform:uppercase">Investasi Fase 1 (aktif)</div>
                  <div style="font-size:18px;font-weight:700;color:#0F6E56;
                              font-family:monospace">Rp {rp(f1)}</div>
                  <div style="font-size:11px;color:#888780">
                    Gym + Penunjang Rp {rp(pen)}<br>
                    Buffer {int(buf_pct*100)}% = Rp {rp((f1+pen)*buf_pct)}
                  </div>
                  <div style="font-size:13px;font-weight:700;color:#0F6E56;
                              margin-top:4px">Total: Rp {rp(xl_data.get("investasi_f1",0))}</div>
                </div>""", unsafe_allow_html=True)
            with col_s2:
                f2_aktif_count  = len([i for i in xl_data.get("fase2_gym",[]) if i["qty"] > 0])
                f2_semua_count  = len(xl_data.get("fase2_gym",[]))
                f2_tambahan     = xl_data.get("investasi_f12", 0) - xl_data.get("investasi_f1", 0)
                if f2_aktif_count > 0:
                    f2_status = f"{f2_aktif_count} item aktif (Qty>0) · {f2_semua_count-f2_aktif_count} belum dibeli"
                    f2_note   = "Masuk ROI jika pilih Fase 1+2"
                else:
                    f2_status = f"{f2_semua_count} item · semua Qty=0"
                    f2_note   = "Belum dihitung dalam ROI Fase 1"
                st.markdown(f"""
                <div style="background:#FDF3E3;border-radius:10px;padding:.75rem 1rem;
                            border-left:4px solid #854F0B">
                  <div style="font-size:11px;color:#888780;font-weight:600;
                              text-transform:uppercase">Fase 2 (pengembangan)</div>
                  <div style="font-size:18px;font-weight:700;color:#854F0B;
                              font-family:monospace">Rp {rp(f2)}</div>
                  <div style="font-size:11px;color:#888780">
                    {f2_status}<br>{f2_note}
                  </div>
                  <div style="font-size:13px;font-weight:700;color:#854F0B;
                              margin-top:4px">Total F1+2: Rp {rp(xl_data.get("investasi_f12",0))}</div>
                </div>""", unsafe_allow_html=True)
            with col_s3:
                st.markdown(f"""
                <div style="background:#EDF3FB;border-radius:10px;padding:.75rem 1rem;
                            border-left:4px solid #185FA5">
                  <div style="font-size:11px;color:#888780;font-weight:600;
                              text-transform:uppercase">Modal Pengelola</div>
                  <div style="font-size:18px;font-weight:700;color:#185FA5;
                              font-family:monospace">Rp {rp(mp)}</div>
                  <div style="font-size:11px;color:#888780">
                    Nilai buku {len(xl_data.get("peralatan",[]))} item peralatan<br>
                    existing per 2026
                  </div>
                  <div style="font-size:13px;font-weight:700;color:#185FA5;
                              margin-top:4px">Aktif dipakai sekarang</div>
                </div>""", unsafe_allow_html=True)

            st.success(
                f"✅ Terbaca.  "
                f"ROI dihitung untuk: **{'Fase 1' if 'Fase 1 saja' in fase_pilihan else 'Fase 1+2'}**  ·  "
                f"Investasi owner: **Rp {rp(inv_aktif)}**  ·  "
                f"Modal pengelola: **Rp {rp(mp)}**"
            )

            # ── Detail tables ─────────────────────────────────────────────────
            tab_f1, tab_f2, tab_pen, tab_alat = st.tabs([
                "🟢 Fase 1 (aktif)", "🟡 Fase 2 (pengembangan)",
                "🏠 Penunjang", "🔧 Alat Pengelola"
            ])

            def show_items(tab, items, label):
                with tab:
                    if items:
                        df = pd.DataFrame(items)
                        df["Qty"] = df["qty"].apply(
                            lambda x: str(x) if x > 0 else "0 — belum dibeli")
                        df["Harga Satuan"] = df["harga"].apply(lambda x: f"Rp {rp(x)}")
                        df["Total"]        = df["total"].apply(lambda x: f"Rp {rp(x)}")
                        display_df = df[["nama","Qty","Harga Satuan","Total"]].rename(
                            columns={"nama": "Item"})
                        st.dataframe(display_df.set_index(
                            pd.RangeIndex(1, len(display_df)+1)),
                            use_container_width=True)
                        # Baris total
                        total_aktif = sum(i["total"] for i in items if i["qty"] > 0)
                        total_semua = sum(i["total"] for i in items)
                        c1, c2 = st.columns(2)
                        with c1:
                            st.caption(f"Subtotal (Qty>0): **Rp {rp(total_aktif)}**")
                        with c2:
                            if total_semua != total_aktif:
                                st.caption(f"Subtotal (semua jika dibeli): **Rp {rp(total_semua)}**")
                    else:
                        st.caption(f"Tidak ada item {label}.")

            show_items(tab_f1,  xl_data.get("fase1_gym",[]),  "Fase 1")
            show_items(tab_f2,  xl_data.get("fase2_gym",[]),  "Fase 2")
            show_items(tab_pen, xl_data.get("pend_items",[]), "penunjang")

            with tab_alat:
                alat = xl_data.get("peralatan",[])
                if alat:
                    df_a = pd.DataFrame(alat)
                    df_a["harga"] = df_a["harga"].apply(lambda x: f"Rp {rp(x)}")
                    df_a["nb"]    = df_a["nb"].apply(lambda x: f"Rp {rp(x)}")
                    st.dataframe(
                        df_a[["nama","thn","harga","nb"]].rename(columns={
                            "nama":"Alat","thn":"Thn Beli",
                            "harga":"Harga Beli","nb":"Nilai Buku 2026"}),
                        use_container_width=True)

        except Exception as e:
            st.error(f"❌ Gagal membaca file: {e}")
            xl_data = {}

# ─── SIDEBAR ─────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Parameter Dasar")

    # ── Investasi & Bagi Hasil ────────────────────────────────────────────────
    st.markdown('<div class="section-header">Investasi &amp; Bagi Hasil</div>', unsafe_allow_html=True)

    investasi = st.number_input("Investasi owner / Pihak I (Rp)",
                                value=xl_data.get("investasi", 101_553_000),
                                step=5_000_000, format="%d",
                                help="Peralatan gym baru + fasilitas penunjang")
    st.markdown(rp_cap(investasi), unsafe_allow_html=True)

    modal_pengelola = st.number_input("Modal pengelola / Pihak II — nilai buku alat (Rp)",
                                      value=xl_data.get("modal_pengelola", 8_757_600),
                                      step=500_000, format="%d",
                                      help="Nilai buku peralatan existing pengelola per 2026")
    st.markdown(rp_cap(modal_pengelola), unsafe_allow_html=True)

    omzet_min = st.number_input("Target omzet minimum BEP (Rp/bln)",
                                value=20_000_000, step=500_000, format="%d")
    st.markdown(rp_cap(omzet_min), unsafe_allow_html=True)

    share_ii = st.slider("Bagi hasil pengelola (%)", 50, 90, 70, 5)
    share_i  = 100 - share_ii
    st.caption(f"Pengelola: **{share_ii}%**  |  Owner: **{share_i}%**")

    # ── Volume Kelas ──────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Volume Kelas / Bulan</div>', unsafe_allow_html=True)
    q_private  = st.number_input("Private 1-on-1 (sesi/bln)",           value=40, min_value=0, step=4)
    q_couple   = st.number_input("Couple / ber-2 (sesi/bln)",            value=16, min_value=0, step=2)
    q_group    = st.number_input("Group 3-4 orang (sesi/bln)",           value=24, min_value=0, step=4)
    q_chair    = st.number_input("Chair exercise (sesi/bln)",             value=20, min_value=0, step=2)
    q_tambahan = st.number_input("Kelas tambahan taichi/yoga (sesi/bln)", value=8,  min_value=0, step=2)
    q_mandiri  = st.number_input("Mandiri tanpa PT (kunjungan/bln)",      value=30, min_value=0, step=5)

    # ── Harga per Sesi ────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Harga per Sesi (Rp)</div>', unsafe_allow_html=True)

    harga_private = st.number_input("Harga — Private 1-on-1 /sesi",
                                    value=195_000, step=5_000, format="%d")
    st.markdown(rp_cap(harga_private), unsafe_allow_html=True)

    harga_couple = st.number_input("Harga — Couple /org/sesi",
                                   value=160_000, step=5_000, format="%d")
    st.markdown(rp_cap(harga_couple), unsafe_allow_html=True)

    harga_group = st.number_input("Harga — Group /org/sesi",
                                  value=110_000, step=5_000, format="%d")
    st.markdown(rp_cap(harga_group), unsafe_allow_html=True)

    harga_chair = st.number_input("Harga — Chair exercise /org/sesi",
                                  value=75_000, step=5_000, format="%d")
    st.markdown(rp_cap(harga_chair), unsafe_allow_html=True)

    harga_tambahan = st.number_input("Harga — Kelas tambahan /org/sesi",
                                     value=100_000, step=5_000, format="%d")
    st.markdown(rp_cap(harga_tambahan), unsafe_allow_html=True)

    harga_mandiri = st.number_input("Harga — Mandiri tanpa PT /kunjungan",
                                    value=50_000, step=5_000, format="%d")
    st.markdown(rp_cap(harga_mandiri), unsafe_allow_html=True)

    peserta_chair    = st.number_input("Peserta chair exercise (org/sesi)", value=8, min_value=1, step=1)
    peserta_tambahan = st.number_input("Peserta kelas tambahan (org/sesi)", value=8, min_value=1, step=1)
    peserta_group    = st.number_input("Peserta group (org/sesi)",          value=3, min_value=1, step=1)
    peserta_couple   = st.number_input("Peserta couple (org/sesi)",         value=2, min_value=1, step=1)

    # ── Biaya Tetap ───────────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Biaya Tetap / Bulan</div>', unsafe_allow_html=True)

    b_trainer = st.number_input("Gaji trainer (Rp)", value=6_000_000, step=500_000, format="%d")
    st.markdown(rp_cap(b_trainer), unsafe_allow_html=True)

    b_admin = st.number_input("Administrasi & ATK (Rp)", value=500_000, step=100_000, format="%d")
    st.markdown(rp_cap(b_admin), unsafe_allow_html=True)

    b_mkt = st.number_input("Marketing & promosi (Rp)", value=1_500_000, step=100_000, format="%d")
    st.markdown(rp_cap(b_mkt), unsafe_allow_html=True)

    b_maint = st.number_input("Maintenance alat (Rp)", value=800_000, step=100_000, format="%d")
    st.markdown(rp_cap(b_maint), unsafe_allow_html=True)

    b_sdm = st.number_input("Pelatihan SDM (Rp)", value=500_000, step=100_000, format="%d")
    st.markdown(rp_cap(b_sdm), unsafe_allow_html=True)

    b_supply = st.number_input("Perlengkapan habis pakai (Rp)", value=400_000, step=50_000, format="%d")
    st.markdown(rp_cap(b_supply), unsafe_allow_html=True)

    b_mgmt = st.number_input("Biaya manajemen operasional (Rp)", value=1_000_000, step=100_000, format="%d")
    st.markdown(rp_cap(b_mgmt), unsafe_allow_html=True)

    # ── Insentif per Kelas ────────────────────────────────────────────────────
    st.markdown('<div class="section-header">Insentif Trainer per Kelas (Rp/sesi)</div>', unsafe_allow_html=True)
    st.caption("Dihitung dari sesi aktual setelah utilisasi")

    ins_private = st.number_input("Insentif — Private 1-on-1",  value=30_000, step=5_000, format="%d")
    st.markdown(rp_cap(ins_private), unsafe_allow_html=True)

    ins_couple = st.number_input("Insentif — Couple / ber-2",   value=25_000, step=5_000, format="%d")
    st.markdown(rp_cap(ins_couple), unsafe_allow_html=True)

    ins_group = st.number_input("Insentif — Group 3-4 orang",   value=20_000, step=5_000, format="%d")
    st.markdown(rp_cap(ins_group), unsafe_allow_html=True)

    ins_chair = st.number_input("Insentif — Chair exercise",     value=20_000, step=5_000, format="%d")
    st.markdown(rp_cap(ins_chair), unsafe_allow_html=True)

    ins_tambahan = st.number_input("Insentif — Kelas tambahan",  value=25_000, step=5_000, format="%d")
    st.markdown(rp_cap(ins_tambahan), unsafe_allow_html=True)

    ins_mandiri = st.number_input("Insentif — Mandiri tanpa PT", value=0,      step=5_000, format="%d")
    st.markdown(rp_cap(ins_mandiri), unsafe_allow_html=True)


    # ── Proyeksi & Asumsi Skenario ────────────────────────────────────────────
    st.markdown('<div class="section-header">Proyeksi</div>', unsafe_allow_html=True)
    proj_years = st.radio("Jangka proyeksi", [3, 5], horizontal=True)

    st.markdown("---")
    st.markdown("### 🎯 Asumsi Skenario")
    st.caption("Sesuaikan utilisasi & pertumbuhan untuk tiap skenario")

    st.markdown("**🔴 Pesimis**")
    util_p   = st.slider("Utilisasi pesimis (%)",      30, 80,  55, 5)
    growth_p = st.slider("Pertumbuhan pesimis (%)",     0, 15,   3, 1)
    cost_p   = st.slider("Kenaikan biaya pesimis (%)", 0, 20,   8, 1)

    st.markdown("**🟡 Moderat**")
    util_m   = st.slider("Utilisasi moderat (%)",      50, 90,  75, 5)
    growth_m = st.slider("Pertumbuhan moderat (%)",     0, 25,  10, 1)
    cost_m   = st.slider("Kenaikan biaya moderat (%)", 0, 15,   5, 1)

    st.markdown("**🟢 Optimis**")
    util_o   = st.slider("Utilisasi optimis (%)",      70, 100, 90, 5)
    growth_o = st.slider("Pertumbuhan optimis (%)",     5,  40, 20, 1)
    cost_o   = st.slider("Kenaikan biaya optimis (%)", 0,  10,  3, 1)


# ─── ARGUMEN UMUM ─────────────────────────────────────────────────────────────
args_common = dict(
    q_private=q_private, q_couple=q_couple, q_group=q_group,
    q_chair=q_chair, q_tambahan=q_tambahan, q_mandiri=q_mandiri,
    harga_private=harga_private, harga_couple=harga_couple, harga_group=harga_group,
    harga_chair=harga_chair, harga_tambahan=harga_tambahan, harga_mandiri=harga_mandiri,
    peserta_couple=peserta_couple, peserta_group=peserta_group,
    peserta_chair=peserta_chair, peserta_tambahan=peserta_tambahan,
    ins_private=ins_private, ins_couple=ins_couple, ins_group=ins_group,
    ins_chair=ins_chair, ins_tambahan=ins_tambahan, ins_mandiri=ins_mandiri,
    b_trainer=b_trainer, b_admin=b_admin, b_mkt=b_mkt, b_maint=b_maint,
    b_sdm=b_sdm, b_supply=b_supply, b_mgmt=b_mgmt,
    investasi=investasi, modal_pengelola=modal_pengelola, omzet_min=omzet_min,
    share_i=share_i, share_ii=share_ii, proj_years=proj_years,
)

d_p = hitung(u_pct=util_p, growth_pct=growth_p, cost_growth_pct=cost_p, **args_common)
d_m = hitung(u_pct=util_m, growth_pct=growth_m, cost_growth_pct=cost_m, **args_common)
d_o = hitung(u_pct=util_o, growth_pct=growth_o, cost_growth_pct=cost_o, **args_common)

skenarios = [
    ("🔴 Pesimis",  d_p, "pesimis",  util_p,  growth_p, cost_p),
    ("🟡 Moderat",  d_m, "moderat",  util_m,  growth_m, cost_m),
    ("🟢 Optimis",  d_o, "optimis",  util_o,  growth_o, cost_o),
]


# ─── TABS ─────────────────────────────────────────────────────────────────────
tab_perbandingan, tab_modal, tab_p, tab_m, tab_o = st.tabs([
    "📊 Perbandingan Skenario",
    "💰 Analisis Modal",
    "🔴 Pesimis",
    "🟡 Moderat",
    "🟢 Optimis",
])


# ════════════════════════════════════════════════════════════════════════════
# TAB PERBANDINGAN
# ════════════════════════════════════════════════════════════════════════════
with tab_perbandingan:
    st.markdown('<div class="section-header">📊 Ringkasan 3 Skenario Side-by-Side</div>', unsafe_allow_html=True)

    col_p, col_m, col_o = st.columns(3)
    for col, (nama, d, css, util, grow, cost) in zip([col_p, col_m, col_o], skenarios):
        with col:
            laba_cls  = "green" if d["laba_bersih"] >= 0 else "red"
            roi_cls   = "green" if d["roi_pct"] >= 15 else "amber" if d["roi_pct"] >= 5 else "red"
            bep_str   = f"BEP bln ke-{d['bep_month']}" if d["bep_month"] else "BEP belum tercapai"
            badge_bep = ('<span class="badge badge-green">✓ Di atas BEP</span>'
                         if d["bep_met"] else '<span class="badge badge-red">✗ Belum BEP</span>')
            st.markdown(f"""
            <div class="skenario-card skenario-{css}">
              <div class="skenario-title">{nama}</div>
              <div class="tab-desc">Utilisasi {util}% · Tumbuh {grow}%/thn · Biaya naik {cost}%/thn</div>
              <table style="width:100%;font-size:13px;border-collapse:collapse">
                <tr><td style="color:#888780;padding:4px 0">Omzet/bln</td>
                    <td style="text-align:right;font-family:DM Mono,monospace;font-weight:600">{fmt_rp(d['total_rev'])}</td></tr>
                <tr><td style="color:#888780;padding:4px 0">Laba bersih/bln</td>
                    <td class="{laba_cls}" style="text-align:right;font-family:DM Mono,monospace;font-weight:600">{fmt_rp(d['laba_bersih'])}</td></tr>
                <tr><td style="color:#888780;padding:4px 0">Bagian pengelola/bln</td>
                    <td class="green" style="text-align:right;font-family:DM Mono,monospace;font-weight:600">{fmt_rp(d['hasil_ii'])}</td></tr>
                <tr><td style="color:#888780;padding:4px 0">Bagian owner/bln</td>
                    <td class="green" style="text-align:right;font-family:DM Mono,monospace;font-weight:600">{fmt_rp(d['hasil_i'])}</td></tr>
                <tr><td style="color:#888780;padding:4px 0">ROI owner/thn</td>
                    <td class="{roi_cls}" style="text-align:right;font-family:DM Mono,monospace;font-weight:600">{d['roi_pct']:.1f}%</td></tr>
                <tr><td style="color:#888780;padding:4px 0">ROI pengelola/thn</td>
                    <td class="blue" style="text-align:right;font-family:DM Mono,monospace;font-weight:600">{d['roi_ii_pct']:.1f}%</td></tr>
                <tr><td style="color:#888780;padding:4px 0">Payback owner</td>
                    <td style="text-align:right;font-size:12px;color:#444441">{bep_str}</td></tr>
                <tr><td style="color:#888780;padding:4px 0">Payback pengelola</td>
                    <td style="text-align:right;font-size:12px;color:#444441">{"bln ke-"+str(d['bep_month_ii']) if d['bep_month_ii'] else "belum tercapai"}</td></tr>
                <tr><td style="color:#888780;padding:4px 0">Status BEP</td>
                    <td style="text-align:right">{badge_bep}</td></tr>
                <tr><td style="color:#888780;padding:4px 0">Kumulatif owner ({proj_years} thn)</td>
                    <td class="{('green' if d['cum_owner'][-1]>=0 else 'red')}" style="text-align:right;font-family:DM Mono,monospace;font-weight:600">{fmt_rp(d['cum_owner'][-1])}</td></tr>
              </table>
            </div>""", unsafe_allow_html=True)

    # Chart perbandingan arus kas kumulatif
    st.markdown('<div class="section-header">📈 Perbandingan Arus Kas Kumulatif Owner</div>', unsafe_allow_html=True)
    months_list = list(range(0, d_m["months"] + 1))
    colors_skenario = {"pesimis": "#E24B4A", "moderat": "#BA7517", "optimis": "#0F6E56"}

    fig_comp = go.Figure()
    for nama, d, css, *_ in skenarios:
        fig_comp.add_trace(go.Scatter(
            x=months_list, y=d["cum_owner"],
            name=nama, line=dict(color=colors_skenario[css], width=2.5),
        ))
    fig_comp.add_hline(y=0, line_dash="dash", line_color="#888780", line_width=1,
                       annotation_text="Break-even", annotation_font=dict(color="#1a1a1a", size=11))
    fig_comp.update_layout(
        paper_bgcolor="white", plot_bgcolor="white", font=CHART_FONT,
        margin=dict(l=16, r=16, t=16, b=48), height=340,
        xaxis=dict(title=dict(text="Bulan", font=TITLE_FONT), tickfont=TICK_FONT,
                   gridcolor="#F0EEE8", linecolor="#D3D1C7"),
        yaxis=dict(title=dict(text="Rp kumulatif (owner)", font=TITLE_FONT), tickfont=TICK_FONT,
                   gridcolor="#F0EEE8", linecolor="#D3D1C7", tickformat=",.0f", tickprefix="Rp "),
        legend=dict(orientation="h", y=-0.2, x=0, font=dict(color="#1a1a1a", size=12)),
        hovermode="x unified"
    )
    st.plotly_chart(fig_comp, use_container_width=True)

    # Bar chart perbandingan omzet & laba
    st.markdown('<div class="section-header">📊 Perbandingan Omzet &amp; Laba Bersih per Skenario</div>', unsafe_allow_html=True)
    fig_bar_comp = go.Figure()
    fig_bar_comp.add_trace(go.Bar(
        name="Omzet/bln", x=["Pesimis","Moderat","Optimis"],
        y=[d_p["total_rev"], d_m["total_rev"], d_o["total_rev"]],
        marker_color=["#FACACA","#F5E3A0","#A8E6CC"],
        text=[fmt_rp(v) for v in [d_p["total_rev"], d_m["total_rev"], d_o["total_rev"]]],
        textposition="outside", textfont=dict(color="#1a1a1a", size=11)
    ))
    fig_bar_comp.add_trace(go.Bar(
        name="Laba bersih/bln", x=["Pesimis","Moderat","Optimis"],
        y=[d_p["laba_bersih"], d_m["laba_bersih"], d_o["laba_bersih"]],
        marker_color=["#E24B4A","#BA7517","#0F6E56"],
        text=[fmt_rp(v) for v in [d_p["laba_bersih"], d_m["laba_bersih"], d_o["laba_bersih"]]],
        textposition="outside", textfont=dict(color="#1a1a1a", size=11)
    ))
    fig_bar_comp.update_layout(
        paper_bgcolor="white", plot_bgcolor="white", font=CHART_FONT,
        barmode="group", margin=dict(l=16, r=16, t=16, b=48), height=320,
        xaxis=dict(tickfont=TICK_FONT, linecolor="#D3D1C7"),
        yaxis=dict(tickfont=TICK_FONT, gridcolor="#F0EEE8", linecolor="#D3D1C7",
                   tickformat=",.0f", tickprefix="Rp "),
        legend=dict(orientation="h", y=-0.2, x=0, font=dict(color="#1a1a1a", size=12)),
    )
    st.plotly_chart(fig_bar_comp, use_container_width=True)

    # Tabel proyeksi tahunan semua skenario
    st.markdown('<div class="section-header">📅 Tabel Proyeksi Tahunan — Semua Skenario</div>', unsafe_allow_html=True)
    yearly_rows = []
    for yr in range(1, proj_years + 1):
        for nama, d, css, util, grow, cost in skenarios:
            rev_yr  = d["total_rev"]  * ((1 + grow / 100) ** (yr - 1)) * 12
            opex_yr = d["total_opex"] * ((1 + cost / 100) ** (yr - 1)) * 12
            laba_yr = rev_yr - opex_yr
            earn_i  = laba_yr * (share_i  / 100) if laba_yr > 0 else 0
            earn_ii = laba_yr * (share_ii / 100) if laba_yr > 0 else 0
            yearly_rows.append({
                "Tahun": f"Tahun {yr}", "Skenario": nama,
                "Omzet (Rp)": fmt_full(rev_yr),
                "Laba Bersih (Rp)": fmt_full(laba_yr),
                f"Bagian Owner {share_i}% (Rp)": fmt_full(earn_i),
                f"Bagian Pengelola {share_ii}% (Rp)": fmt_full(earn_ii),
            })
    df_all = pd.DataFrame(yearly_rows).set_index(["Tahun", "Skenario"])
    st.dataframe(df_all, use_container_width=True)


# ════════════════════════════════════════════════════════════════════════════
# FUNGSI RENDER TAB DETAIL (dipakai 3x)
# ════════════════════════════════════════════════════════════════════════════
def render_tab_detail(d, nama, css, util, grow, cost):
    bep_met     = d["bep_met"]
    laba_bersih = d["laba_bersih"]
    total_rev   = d["total_rev"]
    hasil_ii    = d["hasil_ii"]
    hasil_i     = d["hasil_i"]
    roi_pct     = d["roi_pct"]
    margin      = d["margin"]
    bep_month   = d["bep_month"]
    total_insentif = d["total_insentif"]
    total_opex     = d["total_opex"]
    cum_owner      = d["cum_owner"]
    cum_pengelola  = d["cum_pengelola"]
    months         = d["months"]

    st.markdown(f'<div class="tab-desc">Asumsi: utilisasi <strong>{util}%</strong> · pertumbuhan pendapatan <strong>{grow}%/thn</strong> · kenaikan biaya <strong>{cost}%/thn</strong></div>', unsafe_allow_html=True)

    # Metric cards — baris 1: omzet & laba
    badge_bep = ('<span class="badge badge-green">✓ Di atas BEP</span>'
                 if bep_met else '<span class="badge badge-red">✗ Belum BEP</span>')
    c1, c2, c3 = st.columns(3)
    with c1:
        cls = "green" if laba_bersih >= 0 else "red"
        st.markdown(f"""<div class="metric-card">
          <div class="metric-label">Laba bersih / bulan</div>
          <div class="metric-value {cls}">{fmt_rp(laba_bersih)}</div>
          <div class="metric-sub">Margin {margin:.1f}%  ·  {badge_bep}</div></div>""", unsafe_allow_html=True)
    with c2:
        st.markdown(f"""<div class="metric-card">
          <div class="metric-label">Omzet / bulan</div>
          <div class="metric-value blue">{fmt_rp(total_rev)}</div>
          <div class="metric-sub">Biaya operasional {fmt_rp(total_opex)}</div></div>""", unsafe_allow_html=True)
    with c3:
        st.markdown(f"""<div class="metric-card">
          <div class="metric-label">Bagian pengelola / bulan</div>
          <div class="metric-value green">{fmt_rp(hasil_ii)}</div>
          <div class="metric-sub">Bagian owner {fmt_rp(hasil_i)}/bln</div></div>""", unsafe_allow_html=True)

    # Baris 2: ROI kedua pihak
    c4, c5 = st.columns(2)
    with c4:
        bep_str   = f"Payback owner bln ke-{bep_month}" if bep_month else "Payback owner belum tercapai"
        roi_cls   = "green" if roi_pct >= 15 else "amber" if roi_pct >= 5 else "red"
        st.markdown(f"""<div class="metric-card" style="border-left:4px solid #0F6E56">
          <div class="metric-label">ROI Owner (Pihak I) / tahun</div>
          <div class="metric-value {roi_cls}">{roi_pct:.1f}%</div>
          <div class="metric-sub">{bep_str}  ·  Modal {fmt_rp(d['investasi'])}</div></div>""", unsafe_allow_html=True)
    with c5:
        bep_str_ii  = f"Payback pengelola bln ke-{d['bep_month_ii']}" if d['bep_month_ii'] else "Payback pengelola belum tercapai"
        roi_ii_cls  = "green" if d['roi_ii_pct'] >= 15 else "amber" if d['roi_ii_pct'] >= 5 else "red"
        st.markdown(f"""<div class="metric-card" style="border-left:4px solid #185FA5">
          <div class="metric-label">ROI Pengelola (Pihak II) / tahun</div>
          <div class="metric-value {roi_ii_cls}">{d['roi_ii_pct']:.1f}%</div>
          <div class="metric-sub">{bep_str_ii}  ·  Modal {fmt_rp(d['modal_pengelola'])}</div></div>""", unsafe_allow_html=True)

    # Charts
    col_left, col_right = st.columns([3, 2])
    months_list = list(range(0, months + 1))
    line_color = colors_skenario[css]

    with col_left:
        st.markdown('<div class="section-header">📈 Proyeksi Arus Kas Kumulatif</div>', unsafe_allow_html=True)
        fig_cash = go.Figure()
        fig_cash.add_trace(go.Scatter(x=months_list, y=cum_owner, name="Owner",
            line=dict(color=line_color, width=2.5), fill="tozeroy",
            fillcolor=f"rgba({int(line_color[1:3],16)},{int(line_color[3:5],16)},{int(line_color[5:7],16)},.08)"))
        fig_cash.add_trace(go.Scatter(x=months_list, y=cum_pengelola, name="Pengelola",
            line=dict(color="#185FA5", width=2, dash="dot")))
        fig_cash.add_hline(y=0, line_dash="dash", line_color="#888780", line_width=1,
                           annotation_text="Break-even", annotation_font=dict(color="#1a1a1a", size=11))
        if bep_month:
            fig_cash.add_vline(x=bep_month, line_dash="dot", line_color="#BA7517", line_width=1.5,
                               annotation_text=f"BEP bln {bep_month}", annotation_position="top left",
                               annotation_font=dict(color="#1a1a1a", size=11))
        fig_cash.update_layout(
            paper_bgcolor="white", plot_bgcolor="white", font=CHART_FONT,
            margin=dict(l=16, r=16, t=16, b=48), height=300,
            xaxis=dict(title=dict(text="Bulan", font=TITLE_FONT), tickfont=TICK_FONT,
                       gridcolor="#F0EEE8", linecolor="#D3D1C7"),
            yaxis=dict(title=dict(text="Rp", font=TITLE_FONT), tickfont=TICK_FONT,
                       gridcolor="#F0EEE8", linecolor="#D3D1C7", tickformat=",.0f", tickprefix="Rp "),
            legend=dict(orientation="h", y=-0.22, x=0, font=dict(color="#1a1a1a", size=11)),
            hovermode="x unified")
        st.plotly_chart(fig_cash, use_container_width=True)

    with col_right:
        st.markdown('<div class="section-header">🎯 Pendapatan per Kategori</div>', unsafe_allow_html=True)
        rev_chart_data = [
            ("Private",   d["rev_private"]), ("Couple",  d["rev_couple"]),
            ("Group",     d["rev_group"]),   ("Chair",   d["rev_chair"]),
            ("Tambahan",  d["rev_tambahan"]),("Mandiri", d["rev_mandiri"]),
        ]
        df_rev = (pd.DataFrame(rev_chart_data, columns=["Kategori", "Pendapatan"])
                    .sort_values("Pendapatan", ascending=True))
        fig_bar = go.Figure(go.Bar(
            x=df_rev["Pendapatan"], y=df_rev["Kategori"], orientation="h",
            marker_color=["#888780","#534AB7","#1D9E75","#BA7517","#0F6E56","#185FA5"],
            text=[fmt_rp(v) for v in df_rev["Pendapatan"]], textposition="outside",
            textfont=dict(size=11, color="#1a1a1a")))
        fig_bar.update_layout(
            paper_bgcolor="white", plot_bgcolor="white", font=CHART_FONT,
            margin=dict(l=16, r=90, t=16, b=16), height=300,
            xaxis=dict(showgrid=False, showticklabels=False, tickfont=TICK_FONT),
            yaxis=dict(gridcolor="#F0EEE8", linecolor="#D3D1C7", tickfont=TICK_FONT),
            showlegend=False)
        st.plotly_chart(fig_bar, use_container_width=True)

    # Waterfall
    st.markdown('<div class="section-header">🌊 Waterfall: Omzet → Laba Bersih</div>', unsafe_allow_html=True)
    wf_labels = ["Omzet","Gaji trainer","Admin","Marketing","Maintenance",
                 "Pelatihan SDM","Perlengkapan","Manajemen",
                 "Insentif\nPrivate","Insentif\nCouple","Insentif\nGroup",
                 "Insentif\nChair","Insentif\nTambahan","Insentif\nMandiri","Laba Bersih"]
    wf_values = [total_rev,
                 -d["b_trainer"],-d["b_admin"],-d["b_mkt"],-d["b_maint"],
                 -d["b_sdm"],-d["b_supply"],-d["b_mgmt"],
                 -d["ins_tot_private"],-d["ins_tot_couple"],-d["ins_tot_group"],
                 -d["ins_tot_chair"],-d["ins_tot_tambahan"],-d["ins_tot_mandiri"],
                 laba_bersih]
    wf_measure = ["absolute"] + ["relative"] * (len(wf_labels) - 2) + ["total"]
    fig_wf = go.Figure(go.Waterfall(
        orientation="v", measure=wf_measure, x=wf_labels, y=wf_values,
        text=[fmt_rp(abs(v)) for v in wf_values], textposition="outside",
        textfont=dict(size=10, color="#1a1a1a"),
        connector=dict(line=dict(color="#D3D1C7", width=1, dash="dot")),
        increasing=dict(marker_color="#0F6E56"), decreasing=dict(marker_color="#E24B4A"),
        totals=dict(marker_color="#185FA5")))
    fig_wf.update_layout(
        paper_bgcolor="white", plot_bgcolor="white", font=CHART_FONT,
        margin=dict(l=16, r=16, t=24, b=16), height=360,
        xaxis=dict(tickfont=dict(color="#1a1a1a", size=10), linecolor="#D3D1C7", gridcolor="#F0EEE8"),
        yaxis=dict(tickfont=TICK_FONT, gridcolor="#F0EEE8", linecolor="#D3D1C7",
                   tickformat=",.0f", tickprefix="Rp "), showlegend=False)
    st.plotly_chart(fig_wf, use_container_width=True)

    # Tabel proyeksi tahunan
    st.markdown('<div class="section-header">📅 Tabel Proyeksi Tahunan</div>', unsafe_allow_html=True)
    yearly = []
    for yr in range(1, proj_years + 1):
        rev_yr  = total_rev  * ((1 + grow / 100) ** (yr - 1)) * 12
        opex_yr = total_opex * ((1 + cost / 100) ** (yr - 1)) * 12
        ins_yr  = total_insentif * ((1 + cost / 100) ** (yr - 1)) * 12
        laba_yr = rev_yr - opex_yr
        earn_i  = laba_yr * (share_i  / 100) if laba_yr > 0 else 0
        earn_ii = laba_yr * (share_ii / 100) if laba_yr > 0 else 0
        yearly.append({
            "Tahun": f"Tahun {yr}", "Omzet (Rp)": fmt_full(rev_yr),
            "Total Biaya (Rp)": fmt_full(opex_yr), "  — Insentif (Rp)": fmt_full(ins_yr),
            "Laba Bersih (Rp)": fmt_full(laba_yr),
            f"Bagian Owner {share_i}% (Rp)": fmt_full(earn_i),
            f"Bagian Pengelola {share_ii}% (Rp)": fmt_full(earn_ii),
        })
    st.dataframe(pd.DataFrame(yearly).set_index("Tahun"), use_container_width=True)

    # Breakdown
    col_bd1, col_bd2 = st.columns(2)
    with col_bd1:
        st.markdown('<div class="section-header">💰 Breakdown Pendapatan</div>', unsafe_allow_html=True)
        rev_items = [
            ("Private 1-on-1", d["rev_private"]), ("Couple / ber-2", d["rev_couple"]),
            ("Group 3-4 orang", d["rev_group"]),  ("Chair exercise", d["rev_chair"]),
            ("Kelas tambahan",  d["rev_tambahan"]),("Mandiri tanpa PT",d["rev_mandiri"]),
        ]
        rows_rev = "".join(
            f'<tr><td style="color:#444441">{lbl}</td><td class="green">+{fmt_full(val)}</td></tr>'
            for lbl, val in rev_items)
        st.markdown(f"""<table class="breakdown-table">{rows_rev}
          <tr class="total-row"><td>Total omzet</td><td class="green">+{fmt_full(total_rev)}</td></tr>
        </table>""", unsafe_allow_html=True)

    with col_bd2:
        st.markdown('<div class="section-header">📊 Breakdown Biaya &amp; Bagi Hasil</div>', unsafe_allow_html=True)
        fixed_items = [
            ("Gaji trainer", d["b_trainer"]), ("Administrasi", d["b_admin"]),
            ("Marketing", d["b_mkt"]),        ("Maintenance", d["b_maint"]),
            ("Pelatihan SDM", d["b_sdm"]),    ("Perlengkapan", d["b_supply"]),
            ("Manajemen", d["b_mgmt"]),
        ]
        rows_fixed = "".join(
            f'<tr><td style="color:#444441">{lbl}</td><td class="red">-{fmt_full(val)}</td></tr>'
            for lbl, val in fixed_items)
        insentif_sub = [
            ("↳ Private 1-on-1", d["ins_tot_private"]),  ("↳ Couple / ber-2", d["ins_tot_couple"]),
            ("↳ Group 3-4 orang", d["ins_tot_group"]),   ("↳ Chair exercise", d["ins_tot_chair"]),
            ("↳ Kelas tambahan", d["ins_tot_tambahan"]), ("↳ Mandiri tanpa PT", d["ins_tot_mandiri"]),
        ]
        rows_ins_sub = "".join(
            f'<tr class="sub-row"><td>{lbl}</td>'
            f'<td class="red" style="text-align:right;font-size:12px">-{fmt_full(val)}</td></tr>'
            for lbl, val in insentif_sub)
        laba_cls  = "green" if laba_bersih >= 0 else "red"
        laba_sign = "+" if laba_bersih >= 0 else "-"
        st.markdown(f"""<table class="breakdown-table">
          {rows_fixed}
          <tr><td style="color:#444441;font-weight:500">Insentif trainer (per kelas)</td>
              <td class="red">-{fmt_full(total_insentif)}</td></tr>
          {rows_ins_sub}
          <tr class="total-row"><td>Total biaya operasional</td><td class="red">-{fmt_full(total_opex)}</td></tr>
          <tr class="total-row"><td>Laba bersih</td>
              <td class="{laba_cls}">{laba_sign}{fmt_full(abs(laba_bersih))}</td></tr>
          <tr><td style="color:#444441">Bagian pengelola ({share_ii}%)</td>
              <td class="green">+{fmt_full(hasil_ii)}</td></tr>
          <tr><td style="color:#444441">Bagian owner ({share_i}%)</td>
              <td class="green">+{fmt_full(hasil_i)}</td></tr>
        </table>""", unsafe_allow_html=True)

    # Written breakdown
    st.markdown('<div class="section-header">📝 Analisis &amp; Rekomendasi</div>', unsafe_allow_html=True)
    viable      = total_rev >= omzet_min and laba_bersih > 0
    badge_str   = ('<span class="badge badge-green">✅ Layak (Viable)</span>'
                   if viable else '<span class="badge badge-red">⚠️ Perlu Penyesuaian</span>')
    bep_str2    = f"bulan ke-{bep_month}" if bep_month else "belum tercapai dalam rentang proyeksi ini"
    bep_str2_ii = f"bulan ke-{d['bep_month_ii']}" if d['bep_month_ii'] else "belum tercapai dalam rentang proyeksi ini"
    roi_rank    = "sangat menarik" if roi_pct > 20 else "menarik" if roi_pct > 10 else "perlu review"
    roi_ii_rank = "sangat menarik" if d['roi_ii_pct'] > 20 else "menarik" if d['roi_ii_pct'] > 10 else "perlu review"
    margin_rank = "sangat sehat" if margin > 30 else "cukup baik" if margin > 15 else "tipis, perlu optimasi"
    top_rev     = max(rev_items, key=lambda x: x[1])
    ins_pct     = (total_insentif / total_opex * 100) if total_opex > 0 else 0

    st.markdown(f"""
    <div class="written-card">
      <h4>Status Kelayakan {badge_str}</h4>
      <p>Skenario <strong>{nama}</strong> dengan utilisasi <strong>{util}%</strong>
      memproyeksikan omzet bulanan <strong>{fmt_full(total_rev)}</strong>.
      {"Omzet melampaui target minimum BEP <strong>" + fmt_full(omzet_min) + "</strong>, bagi hasil dapat dibagikan." if bep_met else "Omzet belum mencapai target minimum BEP <strong>" + fmt_full(omzet_min) + "</strong>. Bagi hasil belum dapat dibagikan."}</p>

      <h4>Return on Investment — Owner (Pihak I)</h4>
      <p>Modal owner <strong>{fmt_full(d['investasi'])}</strong> ({d['pct_modal_i']:.1f}% total modal kas).
      Payback period tercapai pada <strong>{bep_str2}</strong>.
      ROI tahunan <strong>{roi_pct:.1f}%</strong> tergolong <strong>{roi_rank}</strong>.
      Dalam {proj_years} tahun, kumulatif owner diproyeksikan <strong>{fmt_rp(cum_owner[-1])}</strong>.</p>

      <h4>Return on Investment — Pengelola (Pihak II)</h4>
      <p>Modal pengelola <strong>{fmt_full(d['modal_pengelola'])}</strong> berupa nilai buku peralatan existing
      ({d['pct_modal_ii']:.1f}% total modal kas).
      Payback period tercapai pada <strong>{bep_str2_ii}</strong>.
      ROI tahunan <strong>{d['roi_ii_pct']:.1f}%</strong> tergolong <strong>{roi_ii_rank}</strong> —
      jauh lebih tinggi dari owner karena modal kas pengelola lebih kecil namun porsi bagi hasil lebih besar ({share_ii}%).
      Dalam {proj_years} tahun, kumulatif pengelola diproyeksikan <strong>{fmt_rp(cum_pengelola[-1])}</strong>.</p>

      <h4>Analisis Insentif Trainer</h4>
      <p>Total insentif <strong>{fmt_full(total_insentif)}</strong> ({ins_pct:.1f}% dari biaya operasional).
      Kelas dengan rasio insentif/pendapatan terendah adalah kandidat terbaik untuk di-scale.</p>

      <h4>Rekomendasi</h4>
      <ol style="padding-left:1.2rem;line-height:2">
        <li><strong>Scale chair exercise</strong> — kapasitas besar, insentif kecil per sesi.</li>
        <li><strong>Konversi ke paket 16x</strong> — pendapatan lebih predictable.</li>
        <li><strong>Review insentif per kelas</strong> secara berkala sesuai utilisasi aktual.</li>
        <li><strong>Evaluasi bagi hasil setelah 24 bulan</strong> sesuai klausul term sheet.</li>
      </ol>
    </div>""", unsafe_allow_html=True)


# ─── RENDER 3 TABS DETAIL ────────────────────────────────────────────────────
with tab_modal:
    d_ref = d_m  # gunakan skenario moderat sebagai referensi modal
    st.markdown('<div class="section-header">💰 Kontribusi Modal Kedua Pihak</div>', unsafe_allow_html=True)

    col_i, col_ii = st.columns(2)
    with col_i:
        st.markdown(f"""
        <div class="metric-card" style="border-left:4px solid #0F6E56">
          <div class="metric-label">Pihak I — Owner / Investor</div>
          <div class="metric-value green">{fmt_full(investasi)}</div>
          <div class="metric-sub">Peralatan gym baru + fasilitas penunjang · {d_ref['pct_modal_i']:.1f}% total modal kas</div>
        </div>""", unsafe_allow_html=True)
    with col_ii:
        st.markdown(f"""
        <div class="metric-card" style="border-left:4px solid #185FA5">
          <div class="metric-label">Pihak II — Pengelola</div>
          <div class="metric-value blue">{fmt_full(modal_pengelola)}</div>
          <div class="metric-sub">Nilai buku peralatan existing 2026 · {d_ref['pct_modal_ii']:.1f}% total modal kas</div>
        </div>""", unsafe_allow_html=True)

    # Donut chart porsi modal
    st.markdown('<div class="section-header">🥧 Porsi Modal vs Porsi Bagi Hasil</div>', unsafe_allow_html=True)
    col_c1, col_c2 = st.columns(2)
    with col_c1:
        fig_donut_modal = go.Figure(go.Pie(
            labels=["Owner (Pihak I)", "Pengelola (Pihak II)"],
            values=[investasi, modal_pengelola],
            hole=.55,
            marker_colors=["#0F6E56", "#185FA5"],
            textfont=dict(color="#1a1a1a", size=12),
            textinfo="label+percent",
        ))
        fig_donut_modal.update_layout(
            title=dict(text="Porsi Modal Kas", font=dict(color="#1a1a1a", size=13)),
            paper_bgcolor="white", font=CHART_FONT,
            margin=dict(l=16, r=16, t=40, b=16), height=260,
            showlegend=False,
            annotations=[dict(text=f"{fmt_rp(investasi+modal_pengelola)}<br>total",
                              x=0.5, y=0.5, font=dict(size=11, color="#1a1a1a"),
                              showarrow=False)]
        )
        st.plotly_chart(fig_donut_modal, use_container_width=True)

    with col_c2:
        fig_donut_bagi = go.Figure(go.Pie(
            labels=["Owner (Pihak I)", "Pengelola (Pihak II)"],
            values=[share_i, share_ii],
            hole=.55,
            marker_colors=["#0F6E56", "#185FA5"],
            textfont=dict(color="#1a1a1a", size=12),
            textinfo="label+percent",
        ))
        fig_donut_bagi.update_layout(
            title=dict(text="Porsi Bagi Hasil Laba", font=dict(color="#1a1a1a", size=13)),
            paper_bgcolor="white", font=CHART_FONT,
            margin=dict(l=16, r=16, t=40, b=16), height=260,
            showlegend=False,
            annotations=[dict(text="Laba<br>bersih",
                              x=0.5, y=0.5, font=dict(size=11, color="#1a1a1a"),
                              showarrow=False)]
        )
        st.plotly_chart(fig_donut_bagi, use_container_width=True)

    # ROI comparison per skenario
    st.markdown('<div class="section-header">📈 Perbandingan ROI Owner vs Pengelola (per skenario)</div>', unsafe_allow_html=True)
    roi_data = {
        "Skenario":      ["🔴 Pesimis", "🟡 Moderat", "🟢 Optimis"],
        "ROI Owner":     [d_p["roi_pct"],    d_m["roi_pct"],    d_o["roi_pct"]],
        "ROI Pengelola": [d_p["roi_ii_pct"], d_m["roi_ii_pct"], d_o["roi_ii_pct"]],
    }
    fig_roi = go.Figure()
    fig_roi.add_trace(go.Bar(name="ROI Owner/thn (%)", x=roi_data["Skenario"],
        y=roi_data["ROI Owner"], marker_color=["#FACACA","#F5E3A0","#A8E6CC"],
        text=[f"{v:.1f}%" for v in roi_data["ROI Owner"]], textposition="outside",
        textfont=dict(color="#1a1a1a", size=11)))
    fig_roi.add_trace(go.Bar(name="ROI Pengelola/thn (%)", x=roi_data["Skenario"],
        y=roi_data["ROI Pengelola"], marker_color=["#E24B4A","#BA7517","#0F6E56"],
        text=[f"{v:.1f}%" for v in roi_data["ROI Pengelola"]], textposition="outside",
        textfont=dict(color="#1a1a1a", size=11)))
    fig_roi.update_layout(
        paper_bgcolor="white", plot_bgcolor="white", font=CHART_FONT,
        barmode="group", margin=dict(l=16, r=16, t=16, b=48), height=300,
        xaxis=dict(tickfont=TICK_FONT, linecolor="#D3D1C7"),
        yaxis=dict(tickfont=TICK_FONT, gridcolor="#F0EEE8", linecolor="#D3D1C7",
                   ticksuffix="%", title=dict(text="ROI / tahun (%)", font=TITLE_FONT)),
        legend=dict(orientation="h", y=-0.22, x=0, font=dict(color="#1a1a1a", size=11)),
    )
    st.plotly_chart(fig_roi, use_container_width=True)

    # Arus kas kedua pihak (skenario moderat)
    st.markdown('<div class="section-header">📈 Proyeksi Arus Kas Kumulatif Kedua Pihak — Skenario Moderat</div>', unsafe_allow_html=True)
    months_list_m = list(range(0, d_m["months"] + 1))
    fig_both = go.Figure()
    fig_both.add_trace(go.Scatter(x=months_list_m, y=d_m["cum_owner"],
        name=f"Owner (modal {fmt_rp(investasi)})",
        line=dict(color="#0F6E56", width=2.5), fill="tozeroy", fillcolor="rgba(15,110,86,.07)"))
    fig_both.add_trace(go.Scatter(x=months_list_m, y=d_m["cum_pengelola"],
        name=f"Pengelola (modal {fmt_rp(modal_pengelola)})",
        line=dict(color="#185FA5", width=2.5), fill="tozeroy", fillcolor="rgba(24,95,165,.07)"))
    fig_both.add_hline(y=0, line_dash="dash", line_color="#888780", line_width=1,
                       annotation_text="Break-even", annotation_font=dict(color="#1a1a1a", size=11))
    if d_m["bep_month"]:
        fig_both.add_vline(x=d_m["bep_month"], line_dash="dot", line_color="#0F6E56", line_width=1.5,
            annotation_text=f"BEP owner bln {d_m['bep_month']}", annotation_position="top right",
            annotation_font=dict(color="#0F6E56", size=10))
    if d_m["bep_month_ii"]:
        fig_both.add_vline(x=d_m["bep_month_ii"], line_dash="dot", line_color="#185FA5", line_width=1.5,
            annotation_text=f"BEP pengelola bln {d_m['bep_month_ii']}", annotation_position="top left",
            annotation_font=dict(color="#185FA5", size=10))
    fig_both.update_layout(
        paper_bgcolor="white", plot_bgcolor="white", font=CHART_FONT,
        margin=dict(l=16, r=16, t=16, b=48), height=340,
        xaxis=dict(title=dict(text="Bulan", font=TITLE_FONT), tickfont=TICK_FONT,
                   gridcolor="#F0EEE8", linecolor="#D3D1C7"),
        yaxis=dict(title=dict(text="Rp kumulatif", font=TITLE_FONT), tickfont=TICK_FONT,
                   gridcolor="#F0EEE8", linecolor="#D3D1C7", tickformat=",.0f", tickprefix="Rp "),
        legend=dict(orientation="h", y=-0.2, x=0, font=dict(color="#1a1a1a", size=11)),
        hovermode="x unified")
    st.plotly_chart(fig_both, use_container_width=True)

    # Tabel ringkasan modal
    st.markdown('<div class="section-header">📋 Tabel Ringkasan — Modal, ROI &amp; Payback per Skenario</div>', unsafe_allow_html=True)
    rows_modal = []
    for nama, d, css, util, grow, cost in skenarios:
        rows_modal.append({
            "Skenario": nama,
            "Modal Owner (Rp)":        fmt_full(investasi),
            "Modal Pengelola (Rp)":    fmt_full(modal_pengelola),
            "ROI Owner / thn":         f"{d['roi_pct']:.1f}%",
            "ROI Pengelola / thn":     f"{d['roi_ii_pct']:.1f}%",
            "Payback Owner":           f"Bln ke-{d['bep_month']}" if d['bep_month'] else "Belum tercapai",
            "Payback Pengelola":       f"Bln ke-{d['bep_month_ii']}" if d['bep_month_ii'] else "Belum tercapai",
            f"Kumulatif Owner {proj_years}thn": fmt_full(d['cum_owner'][-1]),
            f"Kumulatif Pengelola {proj_years}thn": fmt_full(d['cum_pengelola'][-1]),
        })
    st.dataframe(pd.DataFrame(rows_modal).set_index("Skenario"), use_container_width=True)

    # Catatan analisis
    st.markdown(f"""
    <div class="written-card" style="margin-top:16px">
      <h4>📌 Analisis Keseimbangan Modal vs Bagi Hasil</h4>
      <p>
        Owner berkontribusi <strong>{fmt_full(investasi)}</strong> ({d_ref['pct_modal_i']:.1f}% modal kas)
        dan mendapat <strong>{share_i}% bagi hasil</strong>.
        Pengelola berkontribusi <strong>{fmt_full(modal_pengelola)}</strong> ({d_ref['pct_modal_ii']:.1f}% modal kas)
        namun mendapat <strong>{share_ii}% bagi hasil</strong>.
      </p>
      <p>
        Ketidakseimbangan ini <strong>dapat dijustifikasi</strong> karena pengelola menyumbangkan kontribusi non-kas bernilai tinggi:
        keahlian Active Aging, kurikulum latihan, klien existing (15 orang), dan manajemen operasional harian —
        yang secara praktis adalah aset tak berwujud (<em>intangible assets</em>) yang tidak mudah dinilai secara kas.
      </p>
      <p>
        <strong>Rekomendasi:</strong> Nilailah kontribusi non-kas pengelola secara eksplisit dalam term sheet
        (misalnya goodwill klien existing dinilai Rp X per klien aktif) agar posisi kedua pihak
        seimbang secara legal dan tidak menimbulkan ambiguitas di masa depan.
      </p>
    </div>""", unsafe_allow_html=True)


with tab_p:
    render_tab_detail(d_p, "Pesimis",  "pesimis", util_p, growth_p, cost_p)
with tab_m:
    render_tab_detail(d_m, "Moderat",  "moderat", util_m, growth_m, cost_m)
with tab_o:
    render_tab_detail(d_o, "Optimis",  "optimis", util_o, growth_o, cost_o)


# ─── TOMBOL CETAK / DOWNLOAD PDF ─────────────────────────────────────────────
st.markdown("---")
st.markdown('<div class="section-header">🖨️ Cetak / Download Laporan PDF</div>',
            unsafe_allow_html=True)

col_dl1, col_dl2 = st.columns([2, 3])
with col_dl1:
    st.markdown("""
    <div style="background:#FFFFFF;border:1px solid #E8E6DE;border-radius:12px;
                padding:1rem 1.25rem;font-size:13px;color:#444441;line-height:1.7">
      <b>Laporan berisi:</b><br>
      📄 Cover &amp; ringkasan eksekutif<br>
      📊 Perbandingan 3 skenario (tabel + chart)<br>
      💰 Struktur modal &amp; porsi bagi hasil<br>
      📈 Proyeksi arus kas kumulatif<br>
      🔍 Breakdown pendapatan &amp; biaya<br>
      📋 ROI &amp; payback period semua skenario<br>
      📝 Analisis &amp; rekomendasi strategis
    </div>""", unsafe_allow_html=True)

with col_dl2:
    st.markdown("""
    <div style="background:#F0FAF5;border:1px solid #A8E6CC;border-radius:12px;
                padding:.75rem 1.25rem;font-size:12px;color:#444441;line-height:1.6;
                margin-bottom:10px">
      ⚡ PDF dihasilkan dari data aktual di sidebar saat ini.<br>
      Ubah asumsi → klik tombol lagi → PDF terupdate otomatis.
    </div>""", unsafe_allow_html=True)

    if st.button("🖨️  Generate & Download Laporan PDF",
                 type="primary", use_container_width=True):
        with st.spinner("Menyusun laporan PDF..."):
            try:
                pdf_params = dict(
                    d_p=d_p, d_m=d_m, d_o=d_o,
                    investasi=investasi,
                    modal_pengelola=modal_pengelola,
                    share_i=share_i, share_ii=share_ii,
                    proj_years=proj_years,
                    omzet_min=omzet_min,
                    growth_m=growth_m, cost_m=cost_m,
                    # Fase info dari Excel (jika sudah diupload)
                    fase_label   = xl_data.get("fase_label",
                                       "Fase 1" if xl_data else "Manual"),
                    sub_fase1    = xl_data.get("sub_fase1",   0),
                    sub_fase2    = xl_data.get("sub_fase2",   0),
                    sub_pend     = xl_data.get("sub_pend",    0),
                    buffer_pct   = xl_data.get("buffer_pct",  0.10),
                    investasi_f12= xl_data.get("investasi_f12", investasi),
                )
                pdf_bytes = generate_pdf(pdf_params)
                tanggal   = __import__('datetime').date.today().strftime('%Y%m%d')
                st.download_button(
                    label="📥  Klik di sini untuk download PDF",
                    data=pdf_bytes,
                    file_name=f"GGAC_ROI_Report_{tanggal}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
                st.success(
                    f"✅ Laporan siap!  "
                    f"({len(pdf_bytes)//1024} KB · 7 halaman · "
                    f"3 skenario · proyeksi {proj_years} tahun)"
                )
            except Exception as e:
                st.error(f"❌ Gagal generate PDF: {e}")

st.markdown("<br>", unsafe_allow_html=True)
st.caption("GGAC ROI Calculator — Berdasarkan Term Sheet Kerjasama Operasional & Pricelist GGAC. Proyeksi bersifat estimasi.")
